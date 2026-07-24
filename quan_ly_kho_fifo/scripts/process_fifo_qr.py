import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
import pandas as pd
from datetime import datetime
import os
import sys
import shutil

sys.stdout.reconfigure(encoding='utf-8')

# Fixed Total Warehouse Capacity (Constant)
FIXED_TOTAL_LOCATIONS = 4032

# DYNAMIC PORTABLE PATHS (Works anywhere on any PC/Drive)
script_dir = os.path.dirname(os.path.abspath(__file__))
base_dir = os.path.dirname(script_dir)

input_path = os.path.join(base_dir, 'CHUAN_BI_DU_LIEU.xlsx')
output_dir = os.path.join(base_dir, 'KET_QUA_XUAT_KHO')

os.makedirs(output_dir, exist_ok=True)

now_str = datetime.now().strftime('%Y%m%d_%H%M%S')
output_path = os.path.join(output_dir, f'PHIEU_XUAT_FIFO_{now_str}.xlsx')
latest_path = os.path.join(output_dir, 'KET_QUA_MOI_NHAT.xlsx')

print(f"📁 Thư mục làm việc: {base_dir}")
print(f"📖 Đang đọc dữ liệu từ: {input_path}")
wb_in = openpyxl.load_workbook(input_path, data_only=True)

sheet_names = wb_in.sheetnames
print("Danh sách các Sheet phát hiện trong file:", sheet_names)

def find_col_idx(headers, candidates):
    for cand in candidates:
        for idx, h in enumerate(headers):
            if cand.lower() == str(h).strip().lower():
                return idx
    for cand in candidates:
        for idx, h in enumerate(headers):
            if cand.lower() in str(h).strip().lower():
                return idx
    return None

def find_sheet_by_keywords(sheet_names, target_keywords, default_idx):
    for kw in target_keywords:
        for s in sheet_names:
            if kw.lower() in s.lower():
                return s
    if default_idx < len(sheet_names):
        return sheet_names[default_idx]
    return sheet_names[0]

sheet1_name = find_sheet_by_keywords(sheet_names, ['nhat', 'nhặt', 'nhu cầu', 'danh sách', 'sheet1'], 0)
remaining_sheets = [s for s in sheet_names if s != sheet1_name]
sheet2_name = find_sheet_by_keywords(remaining_sheets, ['ton', 'tồn', 'kho', 'sheet2'], 0) if remaining_sheets else sheet_names[-1]

print(f"👉 Sheet 1 (Nhu cầu cần nhặt): '{sheet1_name}'")
print(f"👉 Sheet 2 (Danh sách tồn kho): '{sheet2_name}'")

# Read Sheet 1
ws1 = wb_in[sheet1_name]
rows1 = list(ws1.iter_rows(values_only=True))

header1_idx = None
for idx, r in enumerate(rows1):
    if r and any(str(c).strip().lower() in ['mã', 'mã hàng', 'mã mã', 'code', 'mã sp'] for c in r if c is not None):
        header1_idx = idx
        break

if header1_idx is None:
    header1_idx = 0

headers1 = [str(c).strip() if c is not None else '' for c in rows1[header1_idx]]
col_code1 = find_col_idx(headers1, ['Mã', 'Mã hàng', 'Code', 'Mã SP'])
col_name1 = find_col_idx(headers1, ['Tên', 'Tên Mã', 'Tên hàng', 'Tên SP', 'Name'])
col_unit1 = find_col_idx(headers1, ['ĐVT', 'Đơn vị', 'Đơn vị tính', 'Unit'])
col_qty1 = find_col_idx(headers1, ['Số Lượng KHSX', 'Số lượng cần nhặt', 'Số lượng', 'Nhu cầu', 'Qty', 'SL'])

if col_code1 is None: col_code1 = 0
if col_name1 is None: col_name1 = 1
if col_unit1 is None: col_unit1 = 2
if col_qty1 is None: col_qty1 = 3

valid_items1 = []
if len(rows1) > header1_idx + 1:
    for r in rows1[header1_idx+1:]:
        if not r or len(r) <= max(col_code1, col_qty1):
            continue
        raw_code = r[col_code1]
        if raw_code is None or str(raw_code).strip() == '' or 'DANH SÁCH' in str(raw_code) or 'Mã' in str(raw_code):
            continue
        
        raw_name = r[col_name1] if col_name1 < len(r) and r[col_name1] is not None else ''
        raw_unit = r[col_unit1] if col_unit1 < len(r) and r[col_unit1] is not None else ''
        raw_qty = r[col_qty1] if col_qty1 < len(r) and r[col_qty1] is not None else 0
        
        try:
            qty_val = float(raw_qty)
            if qty_val > 0:
                valid_items1.append({
                    'Mã': str(raw_code).strip(),
                    'Tên Mã': str(raw_name).strip(),
                    'Đơn vị': str(raw_unit).strip(),
                    'Số lượng cần nhặt': qty_val
                })
        except (ValueError, TypeError):
            continue

if len(valid_items1) > 0:
    df1 = pd.DataFrame(valid_items1)
    df1['Mã_clean'] = df1['Mã'].astype(str).str.strip()
else:
    df1 = pd.DataFrame(columns=['Mã', 'Tên Mã', 'Đơn vị', 'Số lượng cần nhặt', 'Mã_clean'])

print(f"✅ Đã đọc thành công {len(df1)} mặt hàng cần nhặt từ Sheet 1.")

# Read Sheet 2 (Inventory)
ws2 = wb_in[sheet2_name]
rows2 = list(ws2.iter_rows(values_only=True))

header2_idx = None
for idx, r in enumerate(rows2):
    if r and any(str(c).strip().lower() in ['mã', 'mã hàng', 'code'] for c in r if c is not None):
        header2_idx = idx
        break

if header2_idx is None:
    header2_idx = 0

headers2 = [str(c).strip() if c is not None else '' for c in rows2[header2_idx]]
col_code2 = find_col_idx(headers2, ['Mã', 'Mã hàng', 'Code'])
col_name2 = find_col_idx(headers2, ['Tên Mã', 'Tên hàng', 'Tên SP', 'Name'])
col_qty2 = find_col_idx(headers2, ['Số lượng', 'Số lượng tồn', 'SL', 'Qty'])
col_unit2 = find_col_idx(headers2, ['Đơn vị', 'ĐVT', 'Unit'])
col_pallet2 = find_col_idx(headers2, ['pallet_no', 'Pallet', 'Số pallet'])
col_loc2 = find_col_idx(headers2, ['Vị trí', 'Location', 'Vị trí kho'])
col_batch2 = find_col_idx(headers2, ['batch_no', 'Batch', 'Số lô'])
col_nsx2 = find_col_idx(headers2, ['Nsx', 'NSX', 'Ngày sản xuất', 'Date'])
col_nhap2 = find_col_idx(headers2, ['Ngày nhập kho', 'Ngày nhập', 'Receipt Date'])
col_kho2 = find_col_idx(headers2, ['Kho'])
col_tenkho2 = find_col_idx(headers2, ['Tên Kho'])

valid_stock = []
for r in rows2[header2_idx+1:]:
    if not r or len(r) <= max(col_code2 if col_code2 is not None else 0, col_qty2 if col_qty2 is not None else 0):
        continue
    
    code_val = r[col_code2] if col_code2 is not None and col_code2 < len(r) else None
    if code_val is None or str(code_val).strip() == '' or 'DANH SÁCH' in str(code_val) or 'Mã' in str(code_val):
        continue
        
    qty_val = r[col_qty2] if col_qty2 is not None and col_qty2 < len(r) else 0
    try:
        qty_num = float(qty_val)
    except (ValueError, TypeError):
        qty_num = 0.0
        
    valid_stock.append({
        'Mã': str(code_val).strip(),
        'Tên Mã': str(r[col_name2]).strip() if col_name2 is not None and col_name2 < len(r) and r[col_name2] is not None else '',
        'Số lượng': qty_num,
        'Đơn vị': str(r[col_unit2]).strip() if col_unit2 is not None and col_unit2 < len(r) and r[col_unit2] is not None else '',
        'pallet_no': str(r[col_pallet2]).strip() if col_pallet2 is not None and col_pallet2 < len(r) and r[col_pallet2] is not None else 'N/A',
        'Vị trí': str(r[col_loc2]).strip() if col_loc2 is not None and col_loc2 < len(r) and r[col_loc2] is not None else 'N/A',
        'batch_no': str(r[col_batch2]).strip() if col_batch2 is not None and col_batch2 < len(r) and r[col_batch2] is not None else 'N/A',
        'Nsx': r[col_nsx2] if col_nsx2 is not None and col_nsx2 < len(r) else '',
        'Ngày nhập kho': r[col_nhap2] if col_nhap2 is not None and col_nhap2 < len(r) else '',
        'Kho': str(r[col_kho2]).strip() if col_kho2 is not None and col_kho2 < len(r) and r[col_kho2] is not None else '',
        'Tên Kho': str(r[col_tenkho2]).strip() if col_tenkho2 is not None and col_tenkho2 < len(r) and r[col_tenkho2] is not None else '',
        '_original_row': r
    })

df2 = pd.DataFrame(valid_stock)
df2['Mã_clean'] = df2['Mã'].astype(str).str.strip()

total_initial_inventory = df2['Số lượng'].sum()
total_initial_pallets_col_i = df2['pallet_no'].nunique()
total_initial_locations_col_j = df2['Vị trí'].nunique()

print(f"✅ Đã đọc thành công {len(df2)} dòng tồn kho từ Sheet 2 (Tổng tồn ban đầu: {total_initial_inventory:g}, {total_initial_pallets_col_i} Pallet duy nhất tại Cột I, {total_initial_locations_col_j} Vị trí tại Cột J).")

def parse_date(val):
    if pd.isna(val) or val == '' or val is None:
        return datetime.max
    if isinstance(val, datetime):
        return val
    val_str = str(val).strip()
    try:
        return datetime.strptime(val_str, '%m/%d/%Y')
    except:
        try:
            return pd.to_datetime(val_str)
        except:
            return datetime.max

df2['Nsx_dt'] = df2['Nsx'].apply(parse_date)
df2['Nhap_dt'] = df2['Ngày nhập kho'].apply(parse_date)
df2['pallet_str'] = df2['pallet_no'].astype(str)

df2_stock = df2.copy()
df2_stock['SL_Ton_hien_tai'] = df2_stock['Số lượng'].astype(float)

picking_details = []
summary_list = []
shortage_list = []

if not df1.empty:
    for idx, r1 in df1.iterrows():
        code = r1['Mã_clean']
        name = r1['Tên Mã']
        unit = r1['Đơn vị']
        qty_needed = float(r1['Số lượng cần nhặt'])
        
        stock_mask = (df2_stock['Mã_clean'] == code) & (df2_stock['SL_Ton_hien_tai'] > 0)
        stock_df = df2_stock[stock_mask].copy()
        stock_df = stock_df.sort_values(by=['Nsx_dt', 'Nhap_dt', 'pallet_str'])
        
        total_stock_init = stock_df['SL_Ton_hien_tai'].sum() if not stock_df.empty else 0.0
        remaining_demand = qty_needed
        total_picked = 0.0
        pallets_used = []
        
        first_pallet_row = True
        
        if stock_df.empty:
            status = "Không có trong tồn kho"
            shortage_qty = qty_needed
        else:
            for s_idx in stock_df.index:
                if remaining_demand <= 1e-9:
                    break
                
                avail_qty = float(df2_stock.loc[s_idx, 'SL_Ton_hien_tai'])
                if avail_qty <= 0:
                    continue
                    
                pick_qty = min(remaining_demand, avail_qty)
                
                df2_stock.loc[s_idx, 'SL_Ton_hien_tai'] -= pick_qty
                remaining_in_pallet = df2_stock.loc[s_idx, 'SL_Ton_hien_tai']
                
                remaining_demand -= pick_qty
                total_picked += pick_qty
                
                pick_qty_str = f"{pick_qty:g}"
                pallets_used.append(f"Pallet {df2_stock.loc[s_idx, 'pallet_no']} ({df2_stock.loc[s_idx, 'Vị trí']}): {pick_qty_str} {unit}")
                
                nsx_val = df2_stock.loc[s_idx, 'Nsx']
                if isinstance(nsx_val, datetime):
                    nsx_str = nsx_val.strftime('%d/%m/%Y')
                elif isinstance(nsx_val, str) and nsx_val.strip() != '':
                    nsx_str = nsx_val.strip()
                else:
                    nsx_str = 'N/A'
                    
                nhap_val = df2_stock.loc[s_idx, 'Ngày nhập kho']
                if isinstance(nhap_val, datetime):
                    ngay_nhap_str = nhap_val.strftime('%d/%m/%Y')
                elif isinstance(nhap_val, str) and nhap_val.strip() != '':
                    ngay_nhap_str = nhap_val.strip()
                else:
                    ngay_nhap_str = 'N/A'
                
                picking_details.append({
                    'Mã hàng': code,
                    'Tên hàng': name,
                    'Đơn vị': unit,
                    'Nhu cầu xuất': qty_needed if first_pallet_row else None,
                    'Pallet No': df2_stock.loc[s_idx, 'pallet_no'],
                    'Vị trí kho': df2_stock.loc[s_idx, 'Vị trí'],
                    'Số lô (Batch)': df2_stock.loc[s_idx, 'batch_no'],
                    'NSX': nsx_str,
                    'Ngày nhập kho': ngay_nhap_str,
                    'SL Tồn pallet': avail_qty,
                    'SL Nhặt FIFO': pick_qty,
                    'SL Còn lại pallet': remaining_in_pallet
                })
                
                first_pallet_row = False
                
            if remaining_demand <= 1e-9:
                status = "Đủ hàng xuất FIFO"
                shortage_qty = 0.0
            else:
                status = "Thiếu hàng tồn"
                shortage_qty = remaining_demand
                
        summary_list.append({
            'Mã hàng': code,
            'Tên hàng': name,
            'Đơn vị': unit,
            'Số lượng cần nhặt': qty_needed,
            'Tổng tồn kho': total_stock_init,
            'Đã nhặt FIFO': total_picked,
            'Còn thiếu': shortage_qty,
            'Trạng thái': status,
            'Chỉ định Pallet & Vị trí nhặt FIFO': "; ".join(pallets_used) if pallets_used else "Chưa có trong tồn kho"
        })
        
        if shortage_qty > 0:
            shortage_list.append({
                'Mã hàng': code,
                'Tên hàng': name,
                'Đơn vị': unit,
                'Nhu cầu xuất': qty_needed,
                'Hiện có trong kho': total_stock_init,
                'Số lượng còn thiếu': shortage_qty,
                'Tỷ lệ đáp ứng (%)': (total_picked / qty_needed * 100) if qty_needed > 0 else 0,
                'Ghi chú': "Cần bổ sung hàng nhập kho" if total_stock_init == 0 else "Tồn kho không đủ đáp ứng hết nhu cầu"
            })

if len(summary_list) > 0:
    df_summary = pd.DataFrame(summary_list)
else:
    df_summary = pd.DataFrame(columns=['Mã hàng', 'Tên hàng', 'Đơn vị', 'Số lượng cần nhặt', 'Tổng tồn kho', 'Đã nhặt FIFO', 'Còn thiếu', 'Trạng thái', 'Chỉ định Pallet & Vị trí nhặt FIFO'])

if len(picking_details) > 0:
    df_details = pd.DataFrame(picking_details)
else:
    df_details = pd.DataFrame(columns=['Mã hàng', 'Tên hàng', 'Đơn vị', 'Nhu cầu xuất', 'Pallet No', 'Vị trí kho', 'Số lô (Batch)', 'NSX', 'Ngày nhập kho', 'SL Tồn pallet', 'SL Nhặt FIFO', 'SL Còn lại pallet'])

if len(shortage_list) > 0:
    df_shortage = pd.DataFrame(shortage_list)
else:
    df_shortage = pd.DataFrame(columns=['Mã hàng', 'Tên hàng', 'Đơn vị', 'Nhu cầu xuất', 'Hiện có trong kho', 'Số lượng còn thiếu', 'Tỷ lệ đáp ứng (%)', 'Ghi chú'])

# METRICS CALCULATIONS FOR ALL SHEETS BASED ON FIXED 4032 LOCATIONS
total_items_req = len(df_summary)
items_ok = (df_summary['Trạng thái'] == "Đủ hàng xuất FIFO").sum() if not df_summary.empty else 0
items_short = (df_summary['Trạng thái'] == "Thiếu hàng tồn").sum() if not df_summary.empty else 0
items_no_stock = (df_summary['Trạng thái'] == "Không có trong tồn kho").sum() if not df_summary.empty else 0

total_picking_rows = len(df_details)
unique_locs_picked = df_details['Vị trí kho'].nunique() if not df_details.empty else 0
unique_pallets_picked = df_details['Pallet No'].nunique() if not df_details.empty else 0

total_picked_all = df_summary['Đã nhặt FIFO'].sum() if not df_summary.empty else 0.0
total_remaining_stock_all = total_initial_inventory - total_picked_all

# Remaining inventory stats
df_remaining_stock = df2_stock[df2_stock['SL_Ton_hien_tai'] > 0].copy()
rem_locs_count = df_remaining_stock['Vị trí'].nunique() if not df_remaining_stock.empty else 0
empty_locs_count = FIXED_TOTAL_LOCATIONS - rem_locs_count
occupancy_rate = (rem_locs_count / FIXED_TOTAL_LOCATIONS * 100) if FIXED_TOTAL_LOCATIONS > 0 else 0

rem_pallets_count_col_i = df_remaining_stock['pallet_no'].nunique() if not df_remaining_stock.empty else 0
pallets_fully_emptied_count = total_initial_pallets_col_i - rem_pallets_count_col_i

wb_out = openpyxl.Workbook()

font_title = Font(name='Segoe UI', size=14, bold=True, color='1F4E78')
font_subtitle = Font(name='Segoe UI', size=10, italic=True, color='595959')
font_kpi_subtitle = Font(name='Segoe UI', size=10, bold=True, color='1F4E78')
font_header = Font(name='Segoe UI', size=11, bold=True, color='FFFFFF')
font_body = Font(name='Segoe UI', size=10)
font_body_bold = Font(name='Segoe UI', size=10, bold=True)
font_location_bold = Font(name='Segoe UI', size=10, bold=True, color='1F4E78')
font_total = Font(name='Segoe UI', size=11, bold=True, color='1F4E78')

fill_header_navy = PatternFill(start_color='1F4E78', end_color='1F4E78', fill_type='solid')
fill_header_blue = PatternFill(start_color='2F5597', end_color='2F5597', fill_type='solid')
fill_header_red = PatternFill(start_color='C00000', end_color='C00000', fill_type='solid')
fill_header_green = PatternFill(start_color='276A3C', end_color='276A3C', fill_type='solid')

fill_total_top = PatternFill(start_color='D9E1F2', end_color='D9E1F2', fill_type='solid')
fill_total_bottom = PatternFill(start_color='F2F2F2', end_color='F2F2F2', fill_type='solid')

fill_green = PatternFill(start_color='E2EFDA', end_color='E2EFDA', fill_type='solid')
fill_yellow = PatternFill(start_color='FFF2CC', end_color='FFF2CC', fill_type='solid')
fill_red = PatternFill(start_color='FCE4D6', end_color='FCE4D6', fill_type='solid')

font_green = Font(name='Segoe UI', size=10, color='375623', bold=True)
font_yellow = Font(name='Segoe UI', size=10, color='7F6000', bold=True)
font_red = Font(name='Segoe UI', size=10, color='C65911', bold=True)

thin_border = Border(
    left=Side(style='thin', color='D9D9D9'), right=Side(style='thin', color='D9D9D9'),
    top=Side(style='thin', color='D9D9D9'), bottom=Side(style='thin', color='D9D9D9')
)

top_total_border = Border(
    left=Side(style='thin', color='1F4E78'), right=Side(style='thin', color='1F4E78'),
    top=Side(style='medium', color='1F4E78'), bottom=Side(style='medium', color='1F4E78')
)

bottom_total_border = Border(
    left=Side(style='thin', color='1F4E78'), right=Side(style='thin', color='1F4E78'),
    top=Side(style='thin', color='1F4E78'), bottom=Side(style='double', color='1F4E78')
)

align_center = Alignment(horizontal='center', vertical='center')
align_left = Alignment(horizontal='left', vertical='center')
align_right = Alignment(horizontal='right', vertical='center')
align_header = Alignment(horizontal='center', vertical='center', wrap_text=True)

def is_bold_target_col(col_name):
    name_str = str(col_name).strip().lower()
    bold_keywords = [
        'mã', 'code', 
        'tên', 'name', 
        'vị trí', 'location', 'chỉ định pallet',
        'số lượng', 'sl', 'nhu cầu', 'tồn', 'còn thiếu', 'hiện có'
    ]
    return any(kw in name_str for kw in bold_keywords)

def write_formatted_sheet(ws, title_text, kpi_subtext, headers, df_data, fill_hdr=fill_header_navy, sum_cols=None):
    ws.views.sheetView[0].showGridLines = True
    
    ws.cell(row=1, column=1, value=title_text).font = font_title
    ws.cell(row=2, column=1, value=kpi_subtext).font = font_kpi_subtitle
    
    top_total_row_idx = 3
    ws.row_dimensions[top_total_row_idx].height = 26
    
    header_row_idx = 4
    first_data_row = 5
    last_data_row = header_row_idx + len(df_data)
    
    all_headers = ["STT"] + headers
    data_cols = list(df_data.columns)
    
    # Write Top Total Row Cells
    ws.cell(row=top_total_row_idx, column=1, value="").fill = fill_total_top
    ws.cell(row=top_total_row_idx, column=1).border = top_total_border
    
    cell_top_lbl = ws.cell(row=top_total_row_idx, column=2, value="TỔNG CỘNG")
    cell_top_lbl.font = font_total
    cell_top_lbl.alignment = align_center
    cell_top_lbl.fill = fill_total_top
    cell_top_lbl.border = top_total_border
    
    for c_idx, col_name in enumerate(data_cols, 2):
        if c_idx == 2:
            continue
        cell = ws.cell(row=top_total_row_idx, column=c_idx)
        cell.font = font_total
        cell.fill = fill_total_top
        cell.border = top_total_border
        
        if sum_cols and col_name in sum_cols and len(df_data) > 0:
            col_letter = get_column_letter(c_idx)
            cell.value = f"=SUM({col_letter}{first_data_row}:{col_letter}{last_data_row})"
            cell.number_format = '#,##0.####'
            cell.alignment = align_right
        else:
            cell.value = ""

    for c_idx, h_text in enumerate(all_headers, 1):
        cell = ws.cell(row=header_row_idx, column=c_idx, value=h_text)
        cell.font = font_header
        cell.fill = fill_hdr
        cell.alignment = align_header
        cell.border = thin_border
        
    ws.row_dimensions[header_row_idx].height = 28
    
    for r_idx, (df_idx, row_series) in enumerate(df_data.iterrows(), start=first_data_row):
        ws.row_dimensions[r_idx].height = 22
        
        cell_stt = ws.cell(row=r_idx, column=1, value=r_idx - header_row_idx)
        cell_stt.font = font_body
        cell_stt.alignment = align_center
        cell_stt.border = thin_border
        
        for c_idx, col_name in enumerate(data_cols, 2):
            val = row_series[col_name]
            cell = ws.cell(row=r_idx, column=c_idx)
            
            should_bold = is_bold_target_col(col_name)
            
            if 'vị trí' in str(col_name).lower() or 'location' in str(col_name).lower() or 'chỉ định pallet' in str(col_name).lower():
                cell.font = font_location_bold
            elif should_bold:
                cell.font = font_body_bold
            else:
                cell.font = font_body
            
            if pd.isna(val) or val is None or str(val).strip() == '':
                cell.value = ''
                cell.alignment = align_center if col_name in ['Mã hàng', 'Đơn vị', 'Pallet No', 'Vị trí kho', 'NSX', 'Ngày nhập kho'] else align_left
            elif isinstance(val, (int, float)):
                cell.value = val
                cell.number_format = '#,##0.####' if isinstance(val, float) and val % 1 != 0 else '#,##0'
                cell.alignment = align_right
            else:
                cell.value = str(val) if val is not None else ''
                cell.alignment = align_center if col_name in ['Mã hàng', 'Đơn vị', 'Pallet No', 'Vị trí kho', 'NSX', 'Ngày nhập kho'] else align_left
                
            cell.border = thin_border
            
            if col_name == 'Trạng thái':
                if val == "Đủ hàng xuất FIFO":
                    cell.fill = fill_green
                    cell.font = font_green
                elif val == "Thiếu hàng tồn":
                    cell.fill = fill_yellow
                    cell.font = font_yellow
                else:
                    cell.fill = fill_red
                    cell.font = font_red

    if len(df_data) > 0:
        bottom_total_row_idx = last_data_row + 1
        ws.row_dimensions[bottom_total_row_idx].height = 26
        
        ws.cell(row=bottom_total_row_idx, column=1, value="").fill = fill_total_bottom
        ws.cell(row=bottom_total_row_idx, column=1).border = bottom_total_border
        
        cell_bot_lbl = ws.cell(row=bottom_total_row_idx, column=2, value="TỔNG CỘNG")
        cell_bot_lbl.font = font_total
        cell_bot_lbl.alignment = align_center
        cell_bot_lbl.fill = fill_total_bottom
        cell_bot_lbl.border = bottom_total_border
        
        for c_idx, col_name in enumerate(data_cols, 2):
            if c_idx == 2:
                continue
            cell = ws.cell(row=bottom_total_row_idx, column=c_idx)
            cell.font = font_total
            cell.fill = fill_total_bottom
            cell.border = bottom_total_border
            
            if sum_cols and col_name in sum_cols:
                col_letter = get_column_letter(c_idx)
                cell.value = f"=SUM({col_letter}{first_data_row}:{col_letter}{last_data_row})"
                cell.number_format = '#,##0.####'
                cell.alignment = align_right
            else:
                cell.value = ""

    for col in ws.columns:
        max_len = 0
        col_letter = get_column_letter(col[0].column)
        for cell in col:
            if cell.row < top_total_row_idx:
                continue
            val_str = str(cell.value or '')
            if len(val_str) > max_len:
                max_len = len(val_str)
        ws.column_dimensions[col_letter].width = max(max_len + 4, 12)

# Subtitle KPI texts
kpi_subtext_det = f"Thời gian: {datetime.now().strftime('%d/%m/%Y %H:%M:%S')} | TỔNG SỐ DÒNG NHẶT: {total_picking_rows} DÒNG | SỐ PALLET CẦN LẤY: {unique_pallets_picked} PALLET DUY NHẤT | SỐ VỊ TRÍ KHO: {unique_locs_picked} VỊ TRÍ DUY NHẤT"
kpi_subtext_sum = f"Thời gian: {datetime.now().strftime('%d/%m/%Y %H:%M:%S')} | TỔNG MÃ CẦN NHẶT: {total_items_req} Mã | ĐỦ TỒN: {items_ok} Mã | THIẾU TỒN: {items_short} Mã | KHÔNG CÓ TRONG KHO (0 vị trí): {items_no_stock} Mã"
kpi_subtext_short = f"Thời gian: {datetime.now().strftime('%d/%m/%Y %H:%M:%S')} | TỔNG SỐ MÃ THIẾU/CHƯA CÓ TRONG KHO: {len(df_shortage)} Mã | KHÔNG CÓ VỊ TRÍ NÀO TRONG KHO (0 vị trí): {items_no_stock} Mã"

# Write Sheets 1, 2, 3
ws_det = wb_out.active
ws_det.title = "Chi tiết nhặt hàng FIFO"
headers_det = list(df_details.columns)
sum_det = ['Nhu cầu xuất', 'SL Tồn pallet', 'SL Nhặt FIFO', 'SL Còn lại pallet']
write_formatted_sheet(ws_det, "BẢNG CHI TIẾT PHÂN BỔ NHẶT HÀNG FIFO (THEO PALLET & VỊ TRÍ)", kpi_subtext_det, headers_det, df_details, fill_header_blue, sum_cols=sum_det)

ws_sum = wb_out.create_sheet(title="Tổng hợp xuất kho FIFO")
headers_sum = list(df_summary.columns)
sum_summary = ['Số lượng cần nhặt', 'Tổng tồn kho', 'Đã nhặt FIFO', 'Còn thiếu']
write_formatted_sheet(ws_sum, "BẢNG TỔNG HỢP NHẶT HÀNG FIFO", kpi_subtext_sum, headers_sum, df_summary, fill_header_navy, sum_cols=sum_summary)

ws_short = wb_out.create_sheet(title="Cảnh báo thiếu hàng")
headers_short = list(df_shortage.columns)
sum_short = ['Nhu cầu xuất', 'Hiện có trong kho', 'Số lượng còn thiếu']
write_formatted_sheet(ws_short, "DANH SÁCH MẶT HÀNG THIẾU TỒN KHO CẦN BỔ SUNG", kpi_subtext_short, headers_short, df_shortage, fill_header_red, sum_cols=sum_short)

# SHEET 4 - TỒN KHO CÒN LẠI (CHO ĐƠN BỔ SUNG 2, 3...)
ws_rem = wb_out.create_sheet(title="Tồn kho còn lại (Cho đơn sau)")
ws_rem.views.sheetView[0].showGridLines = True

kpi_subtext_rem = f"Thời gian: {datetime.now().strftime('%d/%m/%Y %H:%M:%S')} | TỔNG CÔNG SUẤT KHO: {FIXED_TOTAL_LOCATIONS:,} VỊ TRÍ | ĐANG LƯU TỒN: {rem_locs_count:,} VỊ TRÍ (ĐẠT {occupancy_rate:.1f}% TỶ LỆ LẤP ĐẦY) | VỊ TRÍ KHO TRỐNG: {empty_locs_count:,} VỊ TRÍ | PALLET CÒN HÀNG: {rem_pallets_count_col_i:,} PALLET"
ws_rem.cell(row=1, column=1, value="BẢNG TỒN KHO CÒN LẠI SAU XUẤT (DÙNG CHO ĐƠN BỔ SUNG TIẾP THEO)").font = font_title
ws_rem.cell(row=2, column=1, value=kpi_subtext_rem).font = font_kpi_subtitle

rem_headers = headers2 if headers2 else ['Ngày', 'Kho', 'Tên Kho', 'Mã', 'Tên Mã', 'Số lượng', 'Đơn vị', 'Quy cách', 'pallet_no', 'Vị trí', 'batch_no', 'hsd', 'cat', 'subcat', 'Tình Trạng', 'notes', 'Ngày nhập kho', 'Nsx']

top_total_row_rem = 3
header_row_rem = 4
first_data_rem = 5
last_data_rem = header_row_rem + len(df_remaining_stock)

# Write TOP TOTAL ROW for Remaining Inventory Sheet
ws_rem.row_dimensions[top_total_row_rem].height = 26
ws_rem.cell(row=top_total_row_rem, column=1, value="").fill = fill_total_top
ws_rem.cell(row=top_total_row_rem, column=1).border = top_total_border

cell_top_rem_lbl = ws_rem.cell(row=top_total_row_rem, column=2, value="TỔNG TỒN CÒN LẠI")
cell_top_rem_lbl.font = font_total
cell_top_rem_lbl.alignment = align_center
cell_top_rem_lbl.fill = fill_total_top
cell_top_rem_lbl.border = top_total_border

for c_idx in range(2, len(rem_headers) + 1):
    if c_idx == 2:
        continue
    cell = ws_rem.cell(row=top_total_row_rem, column=c_idx)
    cell.font = font_total
    cell.fill = fill_total_top
    cell.border = top_total_border
    
    if col_qty2 is not None and (c_idx - 1) == col_qty2 and len(df_remaining_stock) > 0:
        col_letter = get_column_letter(c_idx)
        cell.value = f"=SUM({col_letter}{first_data_rem}:{col_letter}{last_data_rem})"
        cell.number_format = '#,##0.####'
        cell.alignment = align_right
    else:
        cell.value = ""

# Write Header Row
for c_idx, h_text in enumerate(rem_headers, 1):
    cell = ws_rem.cell(row=header_row_rem, column=c_idx, value=h_text)
    cell.font = font_header
    cell.fill = fill_header_green
    cell.alignment = align_header
    cell.border = thin_border
ws_rem.row_dimensions[header_row_rem].height = 28

# Write Data Rows
for r_idx, (idx_s, row_s) in enumerate(df_remaining_stock.iterrows(), start=first_data_rem):
    ws_rem.row_dimensions[r_idx].height = 20
    orig_tuple = row_s['_original_row']
    
    for c_idx, orig_val in enumerate(orig_tuple, 1):
        cell = ws_rem.cell(row=r_idx, column=c_idx)
        
        h_name = rem_headers[c_idx - 1] if (c_idx - 1) < len(rem_headers) else ''
        should_bold = is_bold_target_col(h_name)
        
        if 'vị trí' in str(h_name).lower() or 'location' in str(h_name).lower():
            cell.font = font_location_bold
        elif should_bold:
            cell.font = font_body_bold
        else:
            cell.font = font_body
        
        if col_qty2 is not None and (c_idx - 1) == col_qty2:
            cell.value = row_s['SL_Ton_hien_tai']
            cell.number_format = '#,##0.####'
            cell.alignment = align_right
        else:
            if isinstance(orig_val, (int, float)):
                cell.value = orig_val
                cell.number_format = '#,##0.####' if isinstance(orig_val, float) and orig_val % 1 != 0 else '#,##0'
                cell.alignment = align_right
            elif isinstance(orig_val, datetime):
                cell.value = orig_val.strftime('%Y-%m-%d')
                cell.alignment = align_center
            else:
                cell.value = str(orig_val) if orig_val is not None else ''
                cell.alignment = align_center if c_idx in [1, 2, 4, 7, 8, 9, 10, 11, 12, 17, 18] else align_left
                
        cell.border = thin_border

# Write BOTTOM TOTAL ROW for Remaining Inventory Sheet
if len(df_remaining_stock) > 0:
    bottom_total_row_rem = last_data_rem + 1
    ws_rem.row_dimensions[bottom_total_row_rem].height = 26
    
    ws_rem.cell(row=bottom_total_row_rem, column=1, value="").fill = fill_total_bottom
    ws_rem.cell(row=bottom_total_row_rem, column=1).border = bottom_total_border
    
    cell_bot_rem_lbl = ws_rem.cell(row=bottom_total_row_rem, column=2, value="TỔNG TỒN CÒN LẠI")
    cell_bot_rem_lbl.font = font_total
    cell_bot_rem_lbl.alignment = align_center
    cell_bot_rem_lbl.fill = fill_total_bottom
    cell_bot_rem_lbl.border = bottom_total_border
    
    for c_idx in range(2, len(rem_headers) + 1):
        if c_idx == 2:
            continue
        cell = ws_rem.cell(row=bottom_total_row_rem, column=c_idx)
        cell.font = font_total
        cell.fill = fill_total_bottom
        cell.border = bottom_total_border
        
        if col_qty2 is not None and (c_idx - 1) == col_qty2:
            col_letter = get_column_letter(c_idx)
            cell.value = f"=SUM({col_letter}{first_data_rem}:{col_letter}{last_data_rem})"
            cell.number_format = '#,##0.####'
            cell.alignment = align_right
        else:
            cell.value = ""

for col in ws_rem.columns:
    max_len = 0
    col_letter = get_column_letter(col[0].column)
    for cell in col:
        if cell.row < top_total_row_rem:
            continue
        val_str = str(cell.value or '')
        if len(val_str) > max_len:
            max_len = len(val_str)
    ws_rem.column_dimensions[col_letter].width = max(max_len + 4, 12)

wb_out.save(output_path)

try:
    shutil.copyfile(output_path, latest_path)
except Exception:
    pass

print(f"\n✅ ĐÃ TẠO THÀNH CÔNG BÁO CÁO XUẤT KHO FIFO (ĐÃ CẬP NHẬT ĐƯỜNG DẪN TƯƠNG ĐỐI 100% PORTABLE)!")
print(f"📄 File lưu tại: {output_path}")

try:
    os.startfile(output_path)
except Exception:
    pass
