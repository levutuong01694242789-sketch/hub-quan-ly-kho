import os
import sys
import json
import openpyxl
from datetime import datetime

# Set UTF-8 encoding for standard output
sys.stdout.reconfigure(encoding='utf-8')

OUTPUT_DIR = os.path.dirname(os.path.abspath(__file__))
OUTPUT_JSON_PATH = os.path.join(OUTPUT_DIR, 'financial_data.json')
OUTPUT_JS_PATH = os.path.join(OUTPUT_DIR, 'financial_data.js')
OUTPUT_EXCEL_PATH = os.path.join(OUTPUT_DIR, 'BAO_CAO_TAI_CHINH_CHI_PHI_KHO.xlsx')
TEMPLATE_EXCEL_PATH = os.path.join(OUTPUT_DIR, 'FILE_MAU_NHAP_CHI_PHI_TAI_CHINH_KHO.xlsx')

# Default Reference Dynamic Parameters (Updated: Plastic Steel-Reinforced Pallet Fleet & Damage/Loss Rate)
# 4,038 Pallets @ 1.0M VND/pallet, 5-year depreciation + 3% annual damage/loss rate
DEFAULT_PARAMS = {
    'warehouse_rent_monthly': 180000000.0, # Tiền thuê mặt bằng kho / tháng (VND)
    'total_labor_monthly': 367500000.0,     # Quỹ lương 29 nhân sự kho ca 12h (1QL 45M + 3TK 20M + 25NV 10.5M)
    'staff_count': 29,                       # Tổng 29 nhân sự kho
    'manager_salary': 45000000.0,            # 1 Quản lý kho / Architect: 45 Tr VND/tháng
    'storekeepers_count': 3,                 # 3 Thủ kho
    'storekeeper_salary': 20000000.0,        # Lương 1 Thủ kho: 20 Tr VND/tháng
    'workers_count': 25,                     # 25 Nhân viên kho
    'worker_salary': 10500000.0,             # Lương 1 Nhân viên kho ca 12h: 10.5 Tr VND/tháng
    
    'utilities_monthly': 35000000.0,        # Điện nước & viễn thông kho / tháng (VND)
    'forklift_operating_monthly': 25000000.0,# Nhiên liệu sạc điện & bảo trì xe nâng / tháng (VND)
    'assets_count': 22,                      # Số lượng Xe nâng & Thiết bị IT/Barcode (Cái)
    
    # Pallet Fleet Parameters (Pallet Nhựa Dẻo Có Ống Sắt)
    'pallet_count': 4038,                    # 4,038 Pallet nhựa ống sắt (Khớp 4,032 Ô kệ + 6 Vùng đệm sàn)
    'pallet_unit_price': 1000000.0,          # 1,000,000 VND / 1 Pallet nhựa ống sắt
    'pallet_useful_years': 5,                # Khấu hao Pallet trong 5 năm (60 tháng)
    'pallet_damage_rate_annual': 0.03,       # Tỷ lệ gãy nứt / hư hỏng / hao hụt Pallet (3%/năm)

    'consumables_monthly': 18000000.0,       # Màng PE, băng keo, đai niềng, tem nhãn / tháng (VND)
    'inventory_carrying_rate_annual': 0.15, # Lãi suất cơ hội giam vốn tồn kho (%/năm)
    
    # Capacity & Inventory Cross-Linked Parameters
    'highbay_rack_capacity': 4032,           # 4,032 Ô kệ cao tầng (Dãy A-H)
    'buffer_zones_count': 6,                # 6 Vùng Đệm Sàn (TW, VW, VL)
    'occupied_bins_count': 3798,            # Số ô kệ đang chứa hàng thực tế (Tỷ lệ lấp đầy 94.2%)
    'monthly_handling_volume_kg': 450000.0, # Sản lượng luân chuyển xuất nhập kho (Kg/tháng)
    'monthly_pick_lines_count': 12500,      # Số dòng hàng nhặt (Pick Lines/tháng)
    'inventory_total_value': 35000000000.0  # Tổng giá trị hàng tồn kho thực tế trong kho (VND)
}

def calculate_warehouse_financials(params=None):
    """Calculate Dynamic Financial Metrics with Pallet Fleet Depreciation & Loss Rate."""
    p = dict(DEFAULT_PARAMS)
    if params:
        p.update(params)

    # 0. Pallet Fleet Cost Calculation
    total_pallet_asset_val = p['pallet_count'] * p['pallet_unit_price']
    monthly_pallet_depreciation = total_pallet_asset_val / (p['pallet_useful_years'] * 12)
    monthly_pallet_damage_loss = (total_pallet_asset_val * p['pallet_damage_rate_annual']) / 12
    total_monthly_pallet_cost = monthly_pallet_depreciation + monthly_pallet_damage_loss

    # 1. Total Monthly OPEX (P&L Operating Expense)
    total_facility_cost = p['warehouse_rent_monthly'] + (p['utilities_monthly'] * 0.6)
    total_labor_cost = p['total_labor_monthly']
    total_consumables_cost = p['consumables_monthly']
    total_equipment_cost = p['forklift_operating_monthly'] + (p['utilities_monthly'] * 0.4) + total_monthly_pallet_cost

    total_monthly_opex = total_facility_cost + total_labor_cost + total_consumables_cost + total_equipment_cost
    total_annual_opex = total_monthly_opex * 12

    # 2. Storage & Bin Costs with Real Capacity Occupancy
    total_storage_locations = p['highbay_rack_capacity'] + p['buffer_zones_count']
    occupancy_rate_pct = round((p['occupied_bins_count'] / total_storage_locations) * 100, 1)
    
    cost_per_bin_monthly = round(total_facility_cost / total_storage_locations, 0)
    cost_per_occupied_bin_monthly = round(total_facility_cost / p['occupied_bins_count'], 0)
    cost_per_bin_daily = round(cost_per_bin_monthly / 30, 0)

    # 3. Handling, Labor & Asset Unit Costs
    cost_per_kg_handled = round(total_monthly_opex / p['monthly_handling_volume_kg'], 2)
    cost_per_staff_member = round(total_labor_cost / p['staff_count'], 0) if p['staff_count'] > 0 else 0
    cost_per_asset_device = round(total_equipment_cost / p['assets_count'], 0) if p['assets_count'] > 0 else 0
    cost_per_pick_line = round((total_labor_cost + total_consumables_cost) / p['monthly_pick_lines_count'], 0)

    # Multi-UoM Unit Cost Estimations
    uom_cost_breakdown = {
        'Kg': cost_per_kg_handled,
        'Thùng': round(cost_per_kg_handled * 15, 0),
        'Cái': round(cost_per_kg_handled * 0.5, 0),
        'Cuộn': round(cost_per_kg_handled * 25, 0),
        'Met': round(cost_per_kg_handled * 1.2, 0),
        'Bộ': round(cost_per_kg_handled * 2.0, 0),
        'Đôi': round(cost_per_kg_handled * 0.8, 0),
        'Chai': round(cost_per_kg_handled * 1.0, 0),
        'Vắt': round(cost_per_kg_handled * 0.05, 0)
    }

    # 4. Inventory Carrying Costs
    annual_carrying_cost = round(p['inventory_total_value'] * p['inventory_carrying_rate_annual'], 0)
    monthly_carrying_cost = round(annual_carrying_cost / 12, 0)
    storage_cost_ratio_pct = round((total_facility_cost / p['inventory_total_value']) * 100, 2)

    now_str = datetime.now().strftime("%d/%m/%Y %H:%M:%S")

    financial_summary = {
        'timestamp': now_str,
        'parameters': p,
        'pallet_metrics': {
            'total_pallets': p['pallet_count'],
            'unit_price': p['pallet_unit_price'],
            'total_asset_value': total_pallet_asset_val,
            'monthly_depreciation': monthly_pallet_depreciation,
            'monthly_damage_loss': monthly_pallet_damage_loss,
            'total_monthly_pallet_cost': total_monthly_pallet_cost,
            'cost_per_pallet_monthly': round(total_monthly_pallet_cost / p['pallet_count'], 0)
        },
        'kpi_summary': {
            'total_monthly_opex': total_monthly_opex,
            'total_annual_opex': total_annual_opex,
            'cost_per_bin_monthly': cost_per_bin_monthly,
            'cost_per_occupied_bin_monthly': cost_per_occupied_bin_monthly,
            'cost_per_bin_daily': cost_per_bin_daily,
            'cost_per_kg_handled': cost_per_kg_handled,
            'cost_per_staff_member': cost_per_staff_member,
            'cost_per_asset_device': cost_per_asset_device,
            'cost_per_pick_line': cost_per_pick_line,
            'annual_carrying_cost': annual_carrying_cost,
            'monthly_carrying_cost': monthly_carrying_cost,
            'total_monthly_pallet_cost': total_monthly_pallet_cost,
            'storage_cost_ratio_pct': storage_cost_ratio_pct,
            'total_storage_locations': total_storage_locations,
            'occupied_bins_count': p['occupied_bins_count'],
            'occupancy_rate_pct': occupancy_rate_pct
        },
        'cost_groups_breakdown': {
            'group_1_storage': {
                'title': '🏗️ 1. Chi Phí Lưu Kho & Hạ Tầng',
                'amount_monthly': total_facility_cost,
                'pct_of_total': round((total_facility_cost / total_monthly_opex) * 100, 1),
                'details': f'Mặt bằng kho & điện nước hạ tầng. Tính trên {total_storage_locations:,} vị trí (lấp đầy {occupancy_rate_pct}%).'
            },
            'group_2_labor': {
                'title': '⚡ 2. Chi Phí Nhân Công (Cơ Cấu Thực Tế 29 Người)',
                'amount_monthly': total_labor_cost,
                'pct_of_total': round((total_labor_cost / total_monthly_opex) * 100, 1),
                'details': f'Quỹ lương 29 người: 1 Quản lý (45Tr) + 3 Thủ kho (20Tr) + 25 Nhân viên ca 12h (10.5Tr).'
            },
            'group_3_consumables': {
                'title': '📦 3. Chi Phí Vật Tư & PE Quấn Pallet',
                'amount_monthly': total_consumables_cost,
                'pct_of_total': round((total_consumables_cost / total_monthly_opex) * 100, 1),
                'details': 'Bao gồm màng PE quấn Pallet, băng keo, đai niềng, tem nhãn Barcode/QR.'
            },
            'group_4_equipment': {
                'title': '🚜 4. Chi Phí Xe Nâng & Khấu Hao/Hao Hụt Pallet',
                'amount_monthly': total_equipment_cost,
                'pct_of_total': round((total_equipment_cost / total_monthly_opex) * 100, 1),
                'details': f'Bảo trì xe nâng ({p["forklift_operating_monthly"]:,.0f}) + Khấu hao & hao hụt {p["pallet_count"]:,} Pallet nhựa ống sắt ({total_monthly_pallet_cost:,.0f} VND/tháng).'
            },
            'group_5_carrying': {
                'title': '📉 5. Chi Phí Giam Vốn Tồn Kho (Carrying Cost)',
                'amount_monthly': monthly_carrying_cost,
                'pct_of_total': round((monthly_carrying_cost / (total_monthly_opex + monthly_carrying_cost)) * 100, 1),
                'details': f'Chi phí cơ hội giam vốn ({p["inventory_carrying_rate_annual"]*100}%/năm trên {p["inventory_total_value"]/1000000000:.1f} tỷ tồn kho).'
            }
        },
        'uom_cost_breakdown': uom_cost_breakdown
    }

    # Save output files
    with open(OUTPUT_JSON_PATH, 'w', encoding='utf-8') as f:
        json.dump(financial_summary, f, ensure_ascii=False, indent=2)

    with open(OUTPUT_JS_PATH, 'w', encoding='utf-8') as f_js:
        f_js.write('window.DEFAULT_FINANCIAL_DATA = ')
        json.dump(financial_summary, f_js, ensure_ascii=False)
        f_js.write(';')

    generate_excel_financial_report(financial_summary, OUTPUT_EXCEL_PATH)
    generate_excel_template_file(TEMPLATE_EXCEL_PATH)
    return financial_summary

def generate_excel_template_file(output_path):
    """Generate Structured Sample Excel Template File for Staff Data Entry."""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Nạp Chi Phí Kho"

    headers = [
        "Tiền Thuê Mặt Bằng Kho (VND)",
        "Lương Quản Lý Kho (VND)",
        "Số Lượng Thủ Kho (Người)",
        "Lương Thủ Kho (VND/Người)",
        "Số Lượng Nhân Viên Kho Ca 12H (Người)",
        "Lương Nhân Viên Kho (VND/Người)",
        "Chi Phí Điện Nước Kho (VND)",
        "Số Lượng Pallet Nhựa Ống Sắt (Cái)",
        "Giá 1 Pallet Nhựa Ống Sắt (VND)",
        "Tỷ Lệ Hao Hụt/Bể Pallet (%/Năm)",
        "Chi Phí Xe Nâng & Thiết Bị (VND)",
        "Số Lượng Xe Nâng & IT Assets (Cái)",
        "Chi Phí Màng PE & Vật Tư (VND)",
        "Lãi Suất Giam Vốn (%/Năm)",
        "Tổng Giá Trị Tồn Kho Realtime (VND)",
        "Sản Lượng Xuất Nhập (Kg/Tháng)"
    ]
    ws.append(headers)

    sample_row = [
        180000000,
        45000000,
        3,
        20000000,
        25,
        10500000,
        35000000,
        4038,
        1000000,
        3.0,
        25000000,
        22,
        18000000,
        15.0,
        35000000000,
        450000
    ]
    ws.append(sample_row)

    wb.save(output_path)

def generate_excel_financial_report(data, output_path):
    """Generate Excel Financial Report for CFO & Executive Management."""
    wb = openpyxl.Workbook()
    ws_pl = wb.active
    ws_pl.title = "P&L Báo Cáo Chi Phí Kho"

    ws_pl.append(["BÁO CÁO TỔNG HỢP CHI PHÍ VẬN HÀNH KHO & KHẤU HAO/HAO HỤT PALLET NHỰA ỐNG SẮT"])
    ws_pl.append([f"Thời gian lập báo cáo: {data['timestamp']}"])
    ws_pl.append([])

    pal = data['pallet_metrics']
    ws_pl.append(["BẢNG PHÂN TÍCH CHI PHÍ PALLET NHỰA ỐNG SẮT (1.000.000 VND/CÁI)", "CHỈ SỐ", "GIÁ TRỊ THÁNG (VND)", "DIỄN GIẢI GIẢI TRÌNH"])
    ws_pl.append(["1. Tổng số lượng Pallet nhựa dẻo ống sắt trong kho", f"{pal['total_pallets']:,} Pallet", "-", "Phủ 4,032 ô kệ + 6 vùng đệm"])
    ws_pl.append(["2. Tổng giá trị tài sản Pallet (1 Tr VND/Cái)", f"{pal['total_asset_value']:,.0f} VND", "-", "Vốn đầu tư ban đầu Pallet"])
    ws_pl.append(["3. Chi phí Khấu hao Pallet (Khấu hao 5 năm)", "-", f"{pal['monthly_depreciation']:,.0f} VND", "Phân bổ 60 tháng"])
    ws_pl.append(["4. Chi phí Hao hụt / Gãy nứt / Bể Pallet (3%/Năm)", "-", f"{pal['monthly_damage_loss']:,.0f} VND", "Dự phòng tổn thất xe nâng va chạm"])
    ws_pl.append(["TỔNG CHI PHÍ PALLET HÀNG THÁNG", "-", f"{pal['total_monthly_pallet_cost']:,.0f} VND", f"Bình quân {pal['cost_per_pallet_monthly']:,.0f} VND/Pallet/tháng"])
    ws_pl.append([])

    ws_pl.append(["HẠNG MỤC CHI PHÍ VẬN HÀNH KHO", "CHI PHÍ THÁNG (VND)", "% TRÊN TỔNG CHI PHÍ", "GHI CHÚ DIỄN GIẢI"])

    for k, grp in data['cost_groups_breakdown'].items():
        ws_pl.append([
            grp['title'],
            grp['amount_monthly'],
            f"{grp['pct_of_total']}%",
            grp['details']
        ])

    ws_pl.append([])
    kpi = data['kpi_summary']
    ws_pl.append(["TỔNG CHI PHÍ OPEX KHO HÀNG THÁNG", kpi['total_monthly_opex'], "100%", "Tổng chi phí vận hành kho tháng (Bao gồm Pallet)"])
    ws_pl.append(["TỔNG CHI PHÍ OPEX KHO NĂM (DỰ TOÁN)", kpi['total_annual_opex'], "-", "Dự toán chi phí kho cả năm"])
    ws_pl.append([])
    ws_pl.append(["CHI PHÍ TRÊN 1 Ô KỆ / THÁNG (4,038 Ô KỆ)", kpi['cost_per_bin_monthly'], "VND/Ô Kệ", f"Đơn giá ngày: {kpi['cost_per_bin_daily']:,.0f} VND/ngày"])
    ws_pl.append(["CHI PHÍ BÌNH QUÂN / 1 NHÂN SỰ KHO (29 NGƯỜI)", kpi['cost_per_staff_member'], "VND/Người", "Bình quân 12.67 Tr VND/người trên toàn bộ cơ cấu"])
    ws_pl.append(["CHI PHÍ BÌNH QUÂN / 1 THIẾT BỊ IT & XE NÂNG", kpi['cost_per_asset_device'], "VND/Thiết bị", f"Chi phí bảo trì trên {data['parameters']['assets_count']} thiết bị"])
    ws_pl.append(["CHI PHÍ LƯU KHO & XỬ LÝ TRÊN 1 KG", kpi['cost_per_kg_handled'], "VND/Kg", "Đơn giá bóc tách theo Kg"])

    ws_uom = wb.create_sheet(title="Đơn Giá Theo 9 ĐVT")
    ws_uom.append(["ĐƠN VỊ TÍNH (UoM)", "CHI PHÍ PHÂN BỔ KHO (VND/ĐVT)", "QUY ĐỔI THAM CHIẾU"])
    for uom, price in data['uom_cost_breakdown'].items():
        ws_uom.append([uom, price, f"Đơn giá lưu kho & xử lý cho 1 {uom}"])

    wb.save(output_path)

if __name__ == '__main__':
    try:
        calculate_warehouse_financials()
    except Exception as e:
        print(f"❌ Lỗi tính toán tài chính: {e}")
