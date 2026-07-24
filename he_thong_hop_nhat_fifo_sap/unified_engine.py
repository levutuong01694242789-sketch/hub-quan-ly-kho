import os
import sys
import json
import re
import openpyxl
from datetime import datetime

# Set UTF-8 encoding for standard output
sys.stdout.reconfigure(encoding='utf-8')

# Paths
MASTER_BIN_PATH = r'D:\BÁO CÁO NGÀY\1.ưu tiên xem đầu giờ\LCF_Storagebin_VĐT.xlsx'
DEFAULT_INPUT_DATA = r'C:\Users\My ROG\Desktop\2_LOGISTICS_QUAN_LY_KHO\QUAN_LY_KHO_FIFO\CHUAN_BI_DU_LIEU.xlsx'
DEFAULT_FIFO_RESULT = r'C:\Users\My ROG\Desktop\2_LOGISTICS_QUAN_LY_KHO\QUAN_LY_KHO_FIFO\KET_QUA_XUAT_KHO\KET_QUA_MOI_NHAT.xlsx'

OUTPUT_DIR = os.path.dirname(os.path.abspath(__file__))
OUTPUT_EXCEL_PATH = os.path.join(OUTPUT_DIR, 'KET_QUA_HOP_NHAT_FIFO_SAP.xlsx')
OUTPUT_JSON_PATH = os.path.join(OUTPUT_DIR, 'wms_data.json')
OUTPUT_JS_PATH = os.path.join(OUTPUT_DIR, 'data.js')

def normalize_bin_code(code_str):
    """Normalize old storage bin codes like D1-2-4 -> D1-2-04 for exact matching."""
    if not code_str:
        return []
    s = str(code_str).strip().upper()
    m = re.match(r'^([A-Z]\d+)-(\d+)-(\d+)$', s)
    if m:
        rack_pair, level, pos = m.groups()
        pos_padded = f"{int(pos):02d}"
        pos_unpadded = f"{int(pos)}"
        return [f"{rack_pair}-{level}-{pos_padded}", f"{rack_pair}-{level}-{pos_unpadded}"]
    return [s]

def load_master_bins(bin_file_path=MASTER_BIN_PATH):
    """Load 4,032 storage bin master records from LCF_Storagebin_VĐT.xlsx."""
    if not os.path.exists(bin_file_path):
        raise FileNotFoundError(f"Không tìm thấy file Master Bin: {bin_file_path}")

    wb = openpyxl.load_workbook(bin_file_path, data_only=True)
    bins_db = {}
    old_to_sap = {}

    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        for r in range(4, ws.max_row + 1):
            old_code = ws.cell(row=r, column=4).value # Mã vị trí cũ
            new_code = ws.cell(row=r, column=11).value # Mã vị trí mới (SAP)
            rack = ws.cell(row=r, column=6).value # Dãy (A, B, C, D...)
            level = ws.cell(row=r, column=7).value # Tầng (1, 2, 3, 4)
            pos_num = ws.cell(row=r, column=8).value # Vị trí (01, 02...)
            in_out = ws.cell(row=r, column=9).value # T/N (Trong/Ngoài)
            old_name = ws.cell(row=r, column=10).value # Tên vị trí cũ
            new_name = ws.cell(row=r, column=11).value # Tên vị trí mới
            area = ws.cell(row=r, column=12).value # Khu vực
            sap_wh = ws.cell(row=r, column=13).value # Kho (RM_01_CD, FG_07_CD, AP_01_CD)

            if new_code:
                s_new = str(new_code).strip()
                s_old = str(old_code).strip() if old_code else ""
                
                bin_info = {
                    'sap_code': s_new,
                    'old_code': s_old,
                    'rack': str(rack).strip() if rack else "",
                    'level': int(level) if level is not None else 1,
                    'pos_num': str(pos_num).strip() if pos_num else "",
                    'in_out': str(in_out).strip() if in_out else "",
                    'old_name': str(old_name).strip() if old_name else "",
                    'new_name': str(new_name).strip() if new_name else s_new,
                    'area': str(area).strip() if area else "",
                    'sap_wh': str(sap_wh).strip() if sap_wh else "",
                    'sheet': sheet_name,
                    'bin_type': 'RACK',
                    'pallets': [],
                    'stock_by_unit': {},
                    'status': 'EMPTY'
                }

                bins_db[s_new] = bin_info

                if s_old:
                    old_to_sap[s_old.upper()] = s_new
                    for norm in normalize_bin_code(s_old):
                        old_to_sap[norm.upper()] = s_new

    print(f"✅ Đã nạp thành công {len(bins_db)} Vị trí kho Master chuẩn SAP S/4HANA.")
    return bins_db, old_to_sap

def run_unified_fifo_sap(fifo_file_path=DEFAULT_FIFO_RESULT, master_bin_path=MASTER_BIN_PATH):
    """Run unified FIFO allocation + SAP S/4HANA Storage Bin Mapping & Stock Deduction."""
    bins_db, old_to_sap = load_master_bins(master_bin_path)

    if not os.path.exists(fifo_file_path):
        raise FileNotFoundError(f"Không tìm thấy tệp kết quả FIFO: {fifo_file_path}")

    wb_fifo = openpyxl.load_workbook(fifo_file_path, data_only=True)

    # 1. Read 'Chi tiết nhặt hàng FIFO'
    picked_records = []
    if 'Chi tiết nhặt hàng FIFO' in wb_fifo.sheetnames:
        ws_pick = wb_fifo['Chi tiết nhặt hàng FIFO']
        for r in range(5, ws_pick.max_row + 1):
            stt = ws_pick.cell(row=r, column=1).value
            if not stt:
                continue
            item_code = ws_pick.cell(row=r, column=2).value
            item_name = ws_pick.cell(row=r, column=3).value
            unit = ws_pick.cell(row=r, column=4).value
            qty_req = ws_pick.cell(row=r, column=5).value
            pallet_no = ws_pick.cell(row=r, column=6).value
            old_bin = ws_pick.cell(row=r, column=7).value
            batch_no = ws_pick.cell(row=r, column=8).value
            mfg_date = ws_pick.cell(row=r, column=9).value
            in_date = ws_pick.cell(row=r, column=10).value
            stock_pallet = ws_pick.cell(row=r, column=11).value
            qty_picked = ws_pick.cell(row=r, column=12).value
            qty_rem = ws_pick.cell(row=r, column=13).value

            s_old_bin = str(old_bin).strip() if old_bin else ""
            sap_bin = old_to_sap.get(s_old_bin.upper(), s_old_bin)

            if sap_bin == s_old_bin:
                for norm in normalize_bin_code(s_old_bin):
                    if norm.upper() in old_to_sap:
                        sap_bin = old_to_sap[norm.upper()]
                        break

            picked_records.append({
                'item_code': str(item_code).strip() if item_code else "",
                'item_name': str(item_name).strip() if item_name else "",
                'unit': str(unit).strip() if unit else "Kg",
                'qty_req': float(qty_req) if qty_req else 0.0,
                'pallet_no': str(pallet_no).strip() if pallet_no else "",
                'old_bin': s_old_bin,
                'sap_bin': sap_bin,
                'batch_no': str(batch_no).strip() if batch_no else "",
                'mfg_date': str(mfg_date) if mfg_date else "",
                'qty_picked': float(qty_picked) if qty_picked else 0.0,
                'qty_rem_pallet': float(qty_rem) if qty_rem else 0.0
            })

    # 2. Read 'Tồn kho còn lại (Cho đơn sau)'
    remaining_stock_records = []
    unit_totals_all = {}

    if 'Tồn kho còn lại (Cho đơn sau)' in wb_fifo.sheetnames:
        ws_rem = wb_fifo['Tồn kho còn lại (Cho đơn sau)']
        for r in range(5, ws_rem.max_row + 1):
            sap_wh = ws_rem.cell(row=r, column=2).value
            wh_name = ws_rem.cell(row=r, column=3).value
            item_code = ws_rem.cell(row=r, column=4).value
            item_name = ws_rem.cell(row=r, column=5).value
            qty = ws_rem.cell(row=r, column=6).value
            unit = ws_rem.cell(row=r, column=7).value
            pallet_no = ws_rem.cell(row=r, column=9).value
            old_bin = ws_rem.cell(row=r, column=10).value
            batch_no = ws_rem.cell(row=r, column=11).value
            in_date = ws_rem.cell(row=r, column=17).value
            mfg_date = ws_rem.cell(row=r, column=18).value

            if not item_code and not pallet_no:
                continue

            s_unit = str(unit).strip() if unit else "Kg"
            n_qty = float(qty) if qty else 0.0

            s_old_bin = str(old_bin).strip() if old_bin else ""
            sap_bin = old_to_sap.get(s_old_bin.upper(), s_old_bin)
            if sap_bin == s_old_bin:
                for norm in normalize_bin_code(s_old_bin):
                    if norm.upper() in old_to_sap:
                        sap_bin = old_to_sap[norm.upper()]
                        break

            rec = {
                'sap_wh': str(sap_wh).strip() if sap_wh else "",
                'wh_name': str(wh_name).strip() if wh_name else "",
                'item_code': str(item_code).strip() if item_code else "",
                'item_name': str(item_name).strip() if item_name else "",
                'qty': n_qty,
                'unit': s_unit,
                'pallet_no': str(pallet_no).strip() if pallet_no else "",
                'old_bin': s_old_bin,
                'sap_bin': sap_bin,
                'batch_no': str(batch_no).strip() if batch_no else "",
                'in_date': str(in_date) if in_date else "",
                'mfg_date': str(mfg_date) if mfg_date else ""
            }

            remaining_stock_records.append(rec)
            unit_totals_all[s_unit] = unit_totals_all.get(s_unit, 0.0) + n_qty

            # Handle Buffer / Staging Bins dynamically (TW_, VW_, VL_)
            if sap_bin not in bins_db:
                if any(sap_bin.startswith(prefix) for prefix in ['TW_', 'VW_', 'VL_']):
                    bin_area = "Vùng Đệm Sản Xuất" if sap_bin.startswith('TW_') else ("Vùng Đệm Nguyên Liệu" if sap_bin.startswith('VW_') else "Vùng Đệm Nhập Hàng & Bao Bì")
                    bins_db[sap_bin] = {
                        'sap_code': sap_bin,
                        'old_code': s_old_bin,
                        'rack': 'BUFFER',
                        'level': 0,
                        'pos_num': sap_bin,
                        'in_out': 'S',
                        'old_name': s_old_bin,
                        'new_name': sap_bin,
                        'area': bin_area,
                        'sap_wh': str(sap_wh).strip() if sap_wh else "VÙNG_ĐỆM",
                        'sheet': 'Vùng Đệm Sàn',
                        'bin_type': 'BUFFER',
                        'pallets': [],
                        'stock_by_unit': {},
                        'status': 'FULL'
                    }

            if sap_bin in bins_db:
                bins_db[sap_bin]['pallets'].append(rec)
                bins_db[sap_bin]['stock_by_unit'][s_unit] = bins_db[sap_bin]['stock_by_unit'].get(s_unit, 0.0) + n_qty
                bins_db[sap_bin]['status'] = 'FULL'

    # Filter rack bins (A-H) vs buffer bins
    rack_bins = [b for b in bins_db.values() if b.get('bin_type') == 'RACK']
    buffer_bins = [b for b in bins_db.values() if b.get('bin_type') == 'BUFFER']

    occupied_rack_bins = sum(1 for b in rack_bins if len(b['pallets']) > 0)
    empty_rack_bins = len(rack_bins) - occupied_rack_bins
    occupancy_rate = round((occupied_rack_bins / len(rack_bins)) * 100, 2) if rack_bins else 0.0

    print(f"📊 HỆ THỐNG HỢP NHẤT FIFO & SAP S/4HANA:")
    print(f"  - Kệ Cao Tầng Master SAP (Dãy A-H): {len(rack_bins)} ô")
    print(f"  - Ô Kệ Đang Chứa Hàng: {occupied_rack_bins} ô ({occupancy_rate}%)")
    print(f"  - Ô Kệ Trống Khả Dụng: {empty_rack_bins} ô")
    print(f"  - Vùng Đệm & Trung Chuyển Sàn (TW/VW/VL): {len(buffer_bins)} khu vực")
    print(f"  - Tổng Tồn Kho Theo Đơn Vị Tính: {unit_totals_all}")

    export_json_data = {
        'timestamp': datetime.now().strftime("%d/%m/%Y %H:%M:%S"),
        'summary': {
            'total_bins': len(rack_bins),
            'occupied_bins': occupied_rack_bins,
            'empty_bins': empty_rack_bins,
            'occupancy_rate': occupancy_rate,
            'total_buffer_bins': len(buffer_bins),
            'total_picked_records': len(picked_records),
            'total_remaining_pallets': len(remaining_stock_records),
            'unit_totals_all': unit_totals_all
        },
        'bins': bins_db,
        'picked_records': picked_records
    }

    # Save wms_data.json and data.js
    with open(OUTPUT_JSON_PATH, 'w', encoding='utf-8') as f:
        json.dump(export_json_data, f, ensure_ascii=False, indent=2)

    with open(OUTPUT_JS_PATH, 'w', encoding='utf-8') as f_js:
        f_js.write('window.DEFAULT_WMS_DATA = ')
        json.dump(export_json_data, f_js, ensure_ascii=False)
        f_js.write(';')

    print(f"✅ Đã xuất thành công wms_data.json & data.js cho Web App Hợp Nhất.")

    generate_excel_report(bins_db, picked_records, remaining_stock_records, OUTPUT_EXCEL_PATH)
    return export_json_data

def generate_excel_report(bins_db, picked_records, remaining_stock_records, output_path):
    """Generate Excel report with SAP S/4HANA storage bin codes."""
    wb = openpyxl.Workbook()
    ws_summary = wb.active
    ws_summary.title = "Sơ Đồ Vị Trí Kho SAP"

    headers_summary = ["Mã Vị Trí SAP Mới", "Mã Vị Trí Cũ", "Loại Vị Trí", "Dãy", "Tầng", "Vị Trí", "T/N (Trong/Ngoại)", "Kho SAP", "Khu Vực", "Số Pallet Chứa", "Tồn Kho Theo Đơn Vị Tính", "Trạng Thái Ô"]
    ws_summary.append(headers_summary)

    for bcode, b in bins_db.items():
        stock_str = ", ".join([f"{v:g} {u}" for u, v in b['stock_by_unit'].items()]) if b['stock_by_unit'] else "0"
        b_type_name = "Vùng Đệm Trung Chuyển" if b.get('bin_type') == 'BUFFER' else "Kệ Cao Tầng A-H"
        ws_summary.append([
            b['sap_code'],
            b['old_code'],
            b_type_name,
            b['rack'],
            b['level'],
            b['pos_num'],
            b['in_out'],
            b['sap_wh'],
            b['area'],
            len(b['pallets']),
            stock_str,
            "Có Hàng" if len(b['pallets']) > 0 else "Trống"
        ])

    # Sheet 2: Chi tiết Nhặt FIFO SAP
    ws_pick = wb.create_sheet(title="Chi tiết Nhặt FIFO SAP")
    headers_pick = ["Mã Hàng", "Tên Hàng", "Đơn Vị", "Nhu Cầu Xuất", "Pallet No", "Mã Vị Trí Cũ", "Mã Vị Trí SAP Mới", "Số Lô (Batch)", "SL Nhặt FIFO", "SL Còn Lại Pallet"]
    ws_pick.append(headers_pick)

    for p in picked_records:
        ws_pick.append([
            p['item_code'],
            p['item_name'],
            p['unit'],
            p['qty_req'],
            p['pallet_no'],
            p['old_bin'],
            p['sap_bin'],
            p['batch_no'],
            p['qty_picked'],
            p['qty_rem_pallet']
        ])

    wb.save(output_path)
    print(f"✅ Đã tạo báo cáo Excel hợp nhất: {output_path}")

if __name__ == '__main__':
    try:
        run_unified_fifo_sap()
    except Exception as e:
        print(f"❌ Lỗi xử lý: {e}")
